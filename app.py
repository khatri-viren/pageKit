from flask import Flask, request, render_template, request, redirect, url_for, flash
from openpyxl import Workbook,  load_workbook

app = Flask(__name__)
app.secret_key = 'secretHOHOHO' #Add your own secret key here 

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    form_data = request.form

    try:
        workbook = load_workbook('TestData.xlsx')
    except FileNotFoundError:
        workbook = Workbook()

    # Select active sheet or create a new one
    sheet = workbook.active if workbook.sheetnames else workbook.create_sheet()

    # Append form data to the next available row
    row = sheet.max_row + 1

    # Write form data to Excel
    column = 1
    for field, value in form_data.items():
        # sheet.cell(row=row, column=column, value=field)
        sheet.cell(row=row, column=column, value=value)
        column += 1 

    row += 1
    # Save the workbook
    workbook.save('TestData.xlsx')

    flash('Form submitted successfully', 'success')

    return redirect(url_for('index'))

@app.route('/display_data')
def display_data():
    try:
        workbook = load_workbook('TestData.xlsx')
        sheet = workbook.active

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        return render_template('display.html', data=data)
    except FileNotFoundError:
        flash('No data found in the Excel file', 'danger')
        return render_template('display.html', data=None)


if __name__ == '__main__':
    app.run(debug=True)
